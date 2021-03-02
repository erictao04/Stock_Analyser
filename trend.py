import openpyxl
from openpyxl import styles
import yfinance as yf
import csv
import os
import itertools
from pandas_datareader import data as pdr
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
from collections import OrderedDict


class Trend:
    '''Find probability of price increase after positive and negative trading session'''

    def __init__(self, ticker, after_gain=True, increase=True, auto_append=False, append=True, multi_stocks=False, results_path='Results/results.xlsx'):
        self.after_gain = after_gain
        self.increase = True
        self.ticker = ticker.upper()
        self.append = append
        self.results_path = results_path
        self.auto_append = auto_append
        self.multi_stocks = multi_stocks
        self.error = False

    def get_data(self, ticker=None):
        yf.pdr_override()
        if ticker:
            self.closes = pdr.get_data_yahoo(ticker)['Close']
        else:
            self.closes = pdr.get_data_yahoo(self.ticker)['Close']
        try:
            self.closes[0]
        except:
            self.error = True

    def get_results(self):
        '''Find probabililties'''
        def count(current_close, previous_close, previous_result):
            '''Adds to counts'''
            if current_close > previous_close:
                self.gain_days += 1

                if previous_result == 'gain':
                    self.gain_after_gain += 1
                elif previous_result == 'loss':
                    self.gain_after_loss += 1
                return 'gain'

            elif current_close < previous_close:
                self.loss_days += 1

                return 'loss'

            return

        def count2(current_close, previous_close, previous_result):
            if previous_result == 'gain':
                if current_close > previous_close:
                    self.gain_after_gain += 1
                    self.gain_days += 1
                    return 'gain'
                elif current_close < previous_close:
                    self.loss_days += 1
                    return 'loss'

            elif previous_result == 'loss':
                if current_close > previous_close:
                    self.gain_after_loss += 1
                    self.gain_days += 1
                    return 'gain'
                elif current_close < previous_close:
                    self.loss_days += 1
                    return 'loss'

            else:
                if current_close > previous_close:
                    self.gain_days += 1
                    return 'gain'
                elif current_close < previous_close:
                    self.loss_days += 1
                    return 'loss'

            return

        if self.error:
            return 0, 0, 0, 0, 0, 0

        self.gain_days, self.loss_days, self.gain_after_gain, self.gain_after_loss = 0, 0, 0, 0

        previous_close = self.closes[0]
        previous_result = None

        for (date, close) in self.closes[1:].iteritems():
            previous_result = count(close, previous_close, previous_result)
            previous_close = close
        try:
            self.gain_after_gain_prob = round(100 *
                                              self.gain_after_gain / self.gain_days, 1)
            self.gain_after_loss_prob = round(100 *
                                              self.gain_after_loss / self.loss_days, 1)
        except ZeroDivisionError:
            self.error = True

        if self.multi_stocks:
            return self.gain_after_gain_prob, self.gain_after_loss_prob, self.gain_after_gain, self.gain_days, self.gain_after_loss, self.loss_days

    def export_results(self):
        '''Export results into Excel Sheets'''
        def get_row_num(letter):
            new_row = 4
            for i in itertools.count(start=4):
                if not results_sheet[f'{letter}{i}'].value:
                    new_row = i
                    break
            return new_row

        def setup_sheets(wb):
            '''Set up freeze panes and colouring'''

            def col_width(letters):
                for letter in letters:
                    results_sheet.column_dimensions[letter.upper()].width = 17

            results_sheet = wb['Sheet']
            results_sheet.freeze_panes = 'A4'

            results_sheet.merge_cells('B2:D2')
            results_sheet['B2'] = 'All Stocks'
            results_sheet['B2'].border = Border(left=med, top=med, bottom=med)
            results_sheet['C2'].border = Border(top=med, bottom=med)
            results_sheet['D2'].border = Border(right=med, top=med, bottom=med)
            results_sheet['B3'] = 'Ticker'
            results_sheet['C3'] = 'After Gains (%)'
            results_sheet['D3'] = 'After Loss (%)'
            results_sheet['B3'].border = Border(left=med, bottom=med)
            results_sheet['C3'].border = Border(bottom=med)
            results_sheet['D3'].border = Border(right=med, bottom=med)

            results_sheet.merge_cells('F2:H2')
            results_sheet['F2'].border = Border(left=med, top=med, bottom=med)
            results_sheet['G2'].border = Border(top=med, bottom=med)
            results_sheet['H2'].border = Border(right=med, top=med, bottom=med)
            results_sheet['F2'] = 'All Index Averages'
            results_sheet['F3'] = 'Index'
            results_sheet['G3'] = 'After Gains (%)'
            results_sheet['H3'] = 'After Loss (%)'
            results_sheet['F3'].border = Border(left=med, bottom=med)
            results_sheet['G3'].border = Border(bottom=med)
            results_sheet['H3'].border = Border(right=med, bottom=med)

            results_sheet.merge_cells('J2:L2')
            results_sheet['J2'].border = Border(left=med, top=med, bottom=med)
            results_sheet['K2'].border = Border(top=med, bottom=med)
            results_sheet['L2'].border = Border(right=med, top=med, bottom=med)
            results_sheet['J2'] = 'All Exchange Averages'
            results_sheet['J3'] = 'Exchange'
            results_sheet['K3'] = 'After Gains (%)'
            results_sheet['L3'] = 'After Loss (%)'
            results_sheet['J3'].border = Border(left=med, bottom=med)
            results_sheet['K3'].border = Border(bottom=med)
            results_sheet['L3'].border = Border(right=med, bottom=med)

            for i in range(2, 13):
                letter = get_column_letter(i)
                if letter in ['B', 'F', 'J']:
                    results_sheet[f'{letter}2'].alignment = centered
                results_sheet[f'{letter}3'].alignment = centered

            col_width(['B', 'C', 'D', 'F', 'G', 'H', 'J', 'K', 'L'])

            return results_sheet

        def add_results(sheet, ticker, after_gain_prob, after_loss_prob, new_row, table='All Stocks'):
            '''Add results to Excel Sheets'''
            def add_data(letters, ticker, after_gain_prob, after_loss_prob):
                sheet[f'{letters[0]}{new_row}'] = ticker
                sheet[f'{letters[1]}{new_row}'] = after_gain_prob
                sheet[f'{letters[2]}{new_row}'] = after_loss_prob

                for letter in letters:
                    sheet[f'{letter}{new_row}'].alignment = centered

                sheet[f'{letters[0]}{new_row}'].border = Border(left=med)
                sheet[f'{letters[2]}{new_row}'].border = Border(right=med)

                if new_row % 2 == 0:
                    for letter in letters:
                        sheet[f'{letter}{new_row}'].fill = grey_fill

            if table == 'All Stocks':
                letters = ['B', 'C', 'D']
                if self.multi_stocks:
                    add_data(letters, ticker, after_gain_prob, after_loss_prob)
                else:
                    add_data(letters, self.ticker,
                             self.gain_after_gain_prob, self.gain_after_loss_prob)

            elif table == 'Index Averages':
                letters = ['F', 'G', 'H']
                add_data(letters, ticker, after_gain_prob, after_loss_prob)

            elif table == 'Exchange Averages':
                letters = ['J', 'K', 'L']
                add_data(letters, ticker, after_gain_prob, after_loss_prob)

            return sheet

        centered = Alignment(horizontal='center')
        grey_fill = PatternFill("solid", fgColor='00C0C0C0')
        med = Side(style='medium')

        os.makedirs('Results', exist_ok=True)

        if self.auto_append:
            try:
                results_wb = openpyxl.load_workbook(self.results_path)
                results_sheet = results_wb['Sheet']
            except FileNotFoundError:
                results_wb = openpyxl.Workbook()
                results_sheet = setup_sheets(results_wb)
        elif self.append:
            try:
                results_wb = openpyxl.load_workbook(self.results_path)
                results_sheet = results_wb['Sheet']

            except FileNotFoundError:
                print("File Not Found")
                return
        else:
            results_wb = openpyxl.Workbook()
            results_sheet = setup_sheets(results_wb)

        if self.multi_stocks:
            total_after_gain_prob = round(
                100 * self.total[0] / self.total[1], 1)
            total_after_loss_prob = round(
                100 * self.total[2] / self.total[3], 1)

            new_row = get_row_num('B')

            for (ticker, after_gain_prob, after_loss_prob) in self.all_results:
                add_results(results_sheet, ticker,
                            after_gain_prob, after_loss_prob, new_row, table='All Stocks')
                new_row += 1

            new_row = get_row_num('F')

            for (ticker, (gain_after_gain, gain_days, gain_after_losses, loss_days)) in self.index_results.items():
                try:  # Remove after testing
                    after_gain_prob = round(
                        100 * gain_after_gain / gain_days, 1)
                    after_loss_prob = round(
                        100 * gain_after_losses / loss_days, 1)

                    add_results(results_sheet, ticker.upper(),
                                after_gain_prob, after_loss_prob, new_row, table='Index Averages')
                    new_row += 1

                except ZeroDivisionError:
                    continue

            add_results(results_sheet, 'Total',
                        total_after_gain_prob, total_after_loss_prob, new_row, table='Index Averages')

            new_row = get_row_num('J')

            for (ticker, (gain_after_gain, gain_days, gain_after_losses, loss_days)) in self.exchange_results.items():
                try:  # Remove after testing
                    after_gain_prob = round(
                        100 * gain_after_gain / gain_days, 1)
                    after_loss_prob = round(
                        100 * gain_after_losses / loss_days, 1)

                    add_results(results_sheet, ticker.upper(),
                                after_gain_prob, after_loss_prob, new_row, table='Exchange Averages')
                    new_row += 1
                except ZeroDivisionError:
                    continue

            add_results(results_sheet, 'Total',
                        total_after_gain_prob, total_after_loss_prob, new_row, table='Exchange Averages')

        else:
            if self.error:
                with open('Results/error.txt', 'a') as error_log:
                    error_log.write(
                        f'\nTicker "{self.ticker}" Data Unavailable.')
                self.error = False
                return

            new_row = get_row_num('B')
            results_sheet = add_results(
                results_sheet, None, None, None, new_row)

        results_wb.save(self.results_path)


class MultiTrends(Trend):
    '''Multi stocks analysis using Trend'''

    def __init__(self, tickers):
        self.auto_append = True
        self.multi_stocks = True
        self.error = False
        self.results_path = 'Results/results.xlsx'
        self.tickers = tickers
        self.all_results = []
        self.index_results = OrderedDict()
        self.exchange_results = OrderedDict()

        self.index_results['djia'] = [0, 0, 0, 0]
        self.index_results['sp_500'] = [0, 0, 0, 0]
        self.index_results['nasdaq_100'] = [0, 0, 0, 0]
        self.index_results['russell_2000'] = [0, 0, 0, 0]

        self.exchange_results['nyse_ex'] = [0, 0, 0, 0]
        self.exchange_results['nasdaq_ex'] = [0, 0, 0, 0]
        self.exchange_results['otc_ex'] = [0, 0, 0, 0]

        self.total = [0, 0, 0, 0]

    def get_group_tickers(self):
        def csv_to_list(group, filename):
            lst = []
            with open(f'{group.title()}/{filename}.csv') as csv_file:
                csv_reader = csv.reader(csv_file)
                next(csv_reader)

                for row in csv_reader:
                    lst.append(row[0])

            return lst

        self.djia = csv_to_list('Index', 'DJIA_Tickers')
        self.sp_500 = csv_to_list('Index', 'S&P500_Tickers')
        self.nasdaq_100 = csv_to_list('Index', 'NASDAQ100_Tickers')
        self.russell_2000 = csv_to_list('Index', 'RUSSELL2000_Tickers')

        self.nyse_ex = csv_to_list('Exchange', 'NASDAQ_Ex_Tickers')
        self.nasdaq_ex = csv_to_list('Exchange', 'NYSE_Ex_Tickers')
        self.otc_ex = csv_to_list('Exchange', 'OTC_Ex_Tickers')

    def analyse_stocks(self):
        def add_to_dict(group, gain_after_gain, gain_days, gain_after_loss, loss_days):
            if group in ['djia', 'sp_500', 'nasdaq_100', 'russell_2000']:
                results = self.index_results
            else:
                results = self.exchange_results

            results[group][0] += gain_after_gain
            results[group][1] += gain_days
            results[group][2] += gain_after_loss
            results[group][3] += loss_days

        for ticker in self.tickers:
            self.get_data(ticker=ticker)
            after_gain_prob, after_loss_prob, gain_after_gain, gain_days, gain_after_loss, loss_days = self.get_results()

            if self.error:
                with open('Results/error.txt', 'a') as error_log:
                    error_log.write(f'\nTicker "{ticker}" Data Unavailable.')
                self.error = False
                continue

            if ticker in self.djia:
                add_to_dict('djia', gain_after_gain, gain_days,
                            gain_after_loss, loss_days)

            if ticker in self.sp_500:
                add_to_dict('sp_500', gain_after_gain, gain_days,
                            gain_after_loss, loss_days)

            if ticker in self.nasdaq_100:
                add_to_dict('nasdaq_100', gain_after_gain,
                            gain_days, gain_after_loss, loss_days)

            if ticker in self.russell_2000:
                add_to_dict('russell_2000', gain_after_gain,
                            gain_days, gain_after_loss, loss_days)

            if ticker in self.nyse_ex:
                add_to_dict('nyse_ex', gain_after_gain,
                            gain_days, gain_after_loss, loss_days)

            if ticker in self.nasdaq_ex:
                add_to_dict('nasdaq_ex', gain_after_gain,
                            gain_days, gain_after_loss, loss_days)

            if ticker in self.otc_ex:
                add_to_dict('otc_ex', gain_after_gain, gain_days,
                            gain_after_loss, loss_days)

            self.total[0] += gain_after_gain
            self.total[1] += gain_days
            self.total[2] += gain_after_loss
            self.total[3] += loss_days

            self.all_results.append((ticker, after_gain_prob, after_loss_prob))

        self.export_results()


if __name__ == '__main__':
    with open('All_Tickers.csv') as csv_file:
        csv_reader = csv.reader(csv_file)
        tickers = []

        for ticker in csv_reader:
            tickers.append(ticker[0])

        tickers_obj = MultiTrends(tickers)
        tickers_obj.get_group_tickers()
        tickers_obj.analyse_stocks()

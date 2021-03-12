from trend import MultiTrends, Trend
import yfinance as yf
from pandas_datareader import data as pdr
import openpyxl
from openpyxl.styles import Alignment, PatternFill, table
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
import itertools
import os
import csv


class ConsecutiveDays:
    '''Compare number of consecutive positive or negative days and the average increase in the time frame'''

    def __init__(self, ticker, direction='both', auto_append=False, append=True, multi_stocks=False, results_path='Results/results.xlsx'):
        self.ticker = ticker.upper()
        self.append = append
        self.results_path = results_path
        self.auto_append = auto_append
        self.multi_stocks = multi_stocks
        self.error = False

    def get_data(self, ticker=None):
        yf.pdr_override()
        if ticker:
            self.closes = pdr.get_data_yahoo(
                ticker, start='1990-01-01')['Close']
        else:
            self.closes = pdr.get_data_yahoo(
                self.ticker, start='1990-01-01')['Close']

        try:
            self.closes[0]
        except:
            self.error = True

    def get_results(self, ticker=None):
        if not ticker:
            ticker = self.ticker

        self.results, self.avg_counter, self.avg_change = [], {}, {}
        streak = 0
        previous_close, range_start = self.closes[0], self.closes[0]

        def count(streak, previous_close, close, range_start):
            def append_results(streak, range_start, range_end):
                if streak > 0:
                    avg_change = round(
                        (range_end - range_start) / range_end * 100 / streak, 2)
                else:
                    avg_change = round(
                        (range_start - range_end) / range_start * -100 / streak, 2)

                self.results.append((streak, ticker, avg_change))
                self.avg_counter.setdefault(streak, [0, 0])
                self.avg_counter[streak][0] += 1
                self.avg_counter[streak][1] += avg_change

            if close > previous_close:
                if streak < 0:
                    append_results(streak, range_start, previous_close)
                    range_start = previous_close
                    streak = 0

                streak += 1
            elif close < previous_close:
                if streak > 0:
                    append_results(streak, range_start, previous_close)
                    range_start = previous_close
                    streak = 0

                streak -= 1
            return streak, range_start

        for (date, close) in self.closes[1:].iteritems():
            streak, range_start = count(
                streak, previous_close, close, range_start)
            previous_close = close

        for key in self.avg_counter.keys():
            self.avg_change[key] = round(self.avg_counter[key][1] /
                                         self.avg_counter[key][0], 2)

        if self.multi_stocks:
            return self.results, self.avg_counter

    def export_results(self, results=None, avg_changes=None, ticker=None, avg_count=None):
        '''Exports results into Excel Sheets'''

        centered = Alignment(horizontal='center')
        grey_fill = PatternFill("solid", fgColor='00C0C0C0')
        med = Side(style='medium')

        def get_row_num(letter):
            new_row = 4
            for i in itertools.count(start=4):
                if not results_sheet[f'{letter}{i}'].value:
                    new_row = i
                    break
            return new_row

        def setup_sheets():
            def col_width(letters):
                for letter in letters:
                    sheet.column_dimensions[letter.upper()].width = 17

            if self.auto_append:
                try:
                    wb = openpyxl.load_workbook(self.results_path)
                    sheet = wb['Sheet']
                except FileNotFoundError:
                    os.makedirs('Results', exist_ok=True)
                    wb = openpyxl.Workbook()
            elif self.append:
                try:
                    wb = openpyxl.load_workbook(self.results_path)
                    sheet = wb['Sheet']

                except FileNotFoundError:
                    print("File Not Found")
                    return
            else:
                os.makedirs('Results', exist_ok=True)
                wb = openpyxl.Workbook()

            sheet = wb['Sheet']
            sheet.freeze_panes = 'A4'

            sheet.merge_cells('B2:D2')
            sheet['B2'] = 'All Stocks'
            sheet['B2'].border = Border(left=med, top=med, bottom=med)
            sheet['C2'].border = Border(top=med, bottom=med)
            sheet['D2'].border = Border(right=med, top=med, bottom=med)
            sheet['B3'] = 'Ticker'
            sheet['C3'] = 'Consecutive Days'
            sheet['D3'] = 'Avg Daily Change'
            sheet['B3'].border = Border(left=med, bottom=med)
            sheet['C3'].border = Border(bottom=med)
            sheet['D3'].border = Border(right=med, bottom=med)

            sheet.merge_cells('F2:H2')
            sheet['F2'].border = Border(left=med, top=med, bottom=med)
            sheet['G2'].border = Border(top=med, bottom=med)
            sheet['H2'].border = Border(right=med, top=med, bottom=med)
            sheet['F2'] = 'Average per Consecutive Days'
            sheet['F3'] = 'Index'
            sheet['G3'] = 'Consecutive Days'
            sheet['H3'] = 'Avg Daily Change'
            sheet['F3'].border = Border(left=med, bottom=med)
            sheet['G3'].border = Border(bottom=med)
            sheet['H3'].border = Border(right=med, bottom=med)

            sheet.merge_cells('J2:L2')
            sheet['J2'].border = Border(left=med, top=med, bottom=med)
            sheet['K2'].border = Border(top=med, bottom=med)
            sheet['L2'].border = Border(right=med, top=med, bottom=med)
            sheet['J2'] = 'Frequency Count'
            sheet['J3'] = 'Index'
            sheet['K3'] = 'Consecutive Days'
            sheet['L3'] = 'Frequency'
            sheet['J3'].border = Border(left=med, bottom=med)
            sheet['K3'].border = Border(bottom=med)
            sheet['L3'].border = Border(right=med, bottom=med)

            for i in range(2, 13):
                letter = get_column_letter(i)
                if letter in ['B', 'F', 'J']:
                    sheet[f'{letter}2'].alignment = centered
                sheet[f'{letter}3'].alignment = centered

            col_width(['B', 'C', 'D', 'F', 'G', 'H', 'J', 'K', 'L'])

            return wb, sheet

        def add_results(sheet, ticker, cons_days, avg_change, new_row, table='All Stocks'):
            def add_data(letters, ticker):
                sheet[f'{letters[0]}{new_row}'] = ticker
                sheet[f'{letters[1]}{new_row}'] = cons_days
                sheet[f'{letters[2]}{new_row}'] = avg_change

                for letter in letters:
                    sheet[f'{letter}{new_row}'].alignment = centered

                sheet[f'{letters[0]}{new_row}'].border = Border(left=med)
                sheet[f'{letters[2]}{new_row}'].border = Border(right=med)

                if new_row % 2 == 0:
                    for letter in letters:
                        sheet[f'{letter}{new_row}'].fill = grey_fill

            if table == 'All Stocks':
                letters = ['B', 'C', 'D']
                if self.multi_stocks:
                    add_data(letters, ticker)
                else:
                    add_data(letters, self.ticker)

            elif table == 'Avg Change':
                letters = ['F', 'G', 'H']
                if self.multi_stocks:
                    add_data(letters, ticker.upper())
                else:
                    add_data(letters, self.ticker)

            elif table == 'Avg Count':
                letters = ['J', 'K', 'L']
                if self.multi_stocks:
                    add_data(letters, ticker.upper())
                else:
                    add_data(letters, self.ticker)

            return sheet

        if not results:
            results = self.results

        if not avg_changes:
            avg_changes = self.avg_change

        if not avg_count:
            avg_count = self.avg_counter

        results_wb, results_sheet = setup_sheets()

        new_row = get_row_num('B')
        for cons_days, ticker, avg_change in sorted(results):
            results_sheet = add_results(
                results_sheet, ticker, cons_days, avg_change, new_row)
            new_row += 1

        new_row = get_row_num('F')
        if not self.multi_stocks:
            for cons_days in sorted(avg_changes):
                results_sheet = add_results(
                    results_sheet, self.ticker, cons_days, avg_changes[cons_days], new_row, table='Avg Change')
                new_row += 1
        else:
            for group in avg_changes.keys():
                for cons_days in sorted(avg_changes[group]):
                    results_sheet = add_results(
                        results_sheet, group, cons_days, avg_changes[group][cons_days], new_row, table='Avg Change')
                    new_row += 1

        new_row = get_row_num('J')
        if not self.multi_stocks:
            for cons_days in sorted(avg_count):
                results_sheet = add_results(
                    results_sheet, self.ticker, cons_days, avg_count[cons_days][0], new_row, table='Avg Count')
                new_row += 1
        else:
            for group in avg_count.keys():
                for cons_days in sorted(avg_count[group]):
                    results_sheet = add_results(
                        results_sheet, group, cons_days, avg_count[group][cons_days][0], new_row, table='Avg Count')
                    new_row += 1

        results_wb.save(self.results_path)


class MultiConsecutiveDays(ConsecutiveDays):
    '''Multiple stocks analysis'''

    def __init__(self, tickers, results_path='Results/results.xlsx'):
        self.tickers = tickers
        self.multi_stocks = True
        self.results_path = results_path
        self.auto_append = True

    def get_group_tickers(self):
        def csv_to_list(group, filename):
            lst = []
            with open(f'{group.title()}/{filename}.csv') as csv_file:
                csv_reader = csv.reader(csv_file)

                for row in csv_reader:
                    lst.append(row[0])

            return lst

        self.djia = csv_to_list('Index', 'DJIA_Tickers')
        self.sp_500 = csv_to_list('Index', 'S&P500_Tickers')
        self.nasdaq_100 = csv_to_list('Index', 'NASDAQ100_Tickers')
        self.russell_2000 = csv_to_list('Index', 'RUSSELL2000_Tickers')

        self.nyse_ex = csv_to_list('Exchange', 'NASDAQ_Ex_Tickers')
        self.nasdaq_ex = csv_to_list('Exchange', 'NYSE_Ex_Tickers')
        self.otc_ex = csv_to_list('Exchange', 'OTC_Ex_Tickers')

    def analyse_stocks(self):
        self.all_results = []
        self.all_avg_change, self.all_avg_counter = {}, {}

        def add_to_dict(group, avg_counter):
            results = self.all_avg_counter
            results.setdefault(group, {})

            for (cons_days, avg_counter) in avg_counter.items():
                results[group].setdefault(cons_days, [0, 0])
                results[group][cons_days][0] += avg_counter[0]
                results[group][cons_days][1] += avg_counter[1]

        for ticker in self.tickers:
            self.get_data(ticker=ticker)
            results, avg_counter = self.get_results(ticker=ticker)

            self.all_results += results

            if ticker in self.djia:
                add_to_dict('djia', avg_counter)

        for group in self.all_avg_counter.keys():
            self.all_avg_change.setdefault(group, {})
            for key in self.all_avg_counter[group].keys():
                self.all_avg_change[group][key] = round(self.all_avg_counter[group][key][1] /
                                                        self.all_avg_counter[group][key][0], 2)

        self.export_results(results=self.all_results,
                            avg_changes=self.all_avg_change, avg_count=self.all_avg_counter)


if __name__ == '__main__':
    tickers = []
    with open('Index\\DJIA_Tickers.csv') as csv_file:
        csv_reader = csv.reader(csv_file)

        for ticker in csv_reader:
            tickers.append(ticker[0])

        stock_obj = MultiConsecutiveDays(tickers)
        stock_obj.get_group_tickers()
        stock_obj.analyse_stocks()
